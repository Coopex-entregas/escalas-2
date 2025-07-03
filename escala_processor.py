import pandas as pd
import logging
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)

def get_fill_color(cell):
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.type == "rgb":
        return fill.fgColor.rgb  # Retorna cor ARGB, exemplo: 'FFFF0000' (vermelho)
    return None

def process_escala(file_path):
    try:
        # Lê o arquivo Excel (dados)
        df = pd.read_excel(file_path, header=0)
        logging.info(f"Colunas no arquivo Excel: {df.columns.tolist()}")

        # Abre o Excel com openpyxl para ler as cores
        wb = load_workbook(file_path)
        ws = wb.active  # pega a primeira aba

        colunas_possiveis = {
            'data': ['data', 'DATA', 'Data'],
            'horario': ['horario', 'HORÁRIO', 'horário', 'HORA', 'hora'],
            'contrato': ['contrato', 'CONTRATO', 'Contrato'],
            'nome': ['nome', 'NOME', 'Nome', 'NOME DO COOPERADO', 'Cooperado'],
            'turno': ['turno', 'TURNO', 'Turno']
        }

        def encontrar_nome_coluna(nomes_possiveis, colunas_reais):
            for nome in nomes_possiveis:
                if nome in colunas_reais:
                    return nome
            return None

        colunas_reais = df.columns.tolist()
        coluna_data = encontrar_nome_coluna(colunas_possiveis['data'], colunas_reais)
        coluna_horario = encontrar_nome_coluna(colunas_possiveis['horario'], colunas_reais)
        coluna_contrato = encontrar_nome_coluna(colunas_possiveis['contrato'], colunas_reais)
        coluna_nome = encontrar_nome_coluna(colunas_possiveis['nome'], colunas_reais)
        coluna_turno = encontrar_nome_coluna(colunas_possiveis['turno'], colunas_reais)

        if not coluna_nome:
            logging.error("ERRO: Coluna 'nome' não encontrada.")
            return []

        escala = []
        for idx, row in df.iterrows():
            nome = row.get(coluna_nome, '')
            if pd.notna(nome) and str(nome).strip():
                # No Excel, linha começa na 1, pandas com header=0 linha 0 é linha 2 do Excel
                linha_excel = idx + 2

                # Coluna do nome (índice começa em 0 no pandas, em 1 no openpyxl)
                coluna_idx = df.columns.get_loc(coluna_nome) + 1

                cell = ws.cell(row=linha_excel, column=coluna_idx)
                cor = get_fill_color(cell)  # Pega a cor da célula

                item = {
                    'data': str(row.get(coluna_data, '')) if coluna_data else '',
                    'horario': str(row.get(coluna_horario, '')) if coluna_horario else '',
                    'contrato': str(row.get(coluna_contrato, '')) if coluna_contrato else '',
                    'nome': str(nome),
                    'turno': str(row.get(coluna_turno, '')) if coluna_turno else '',
                    'cor_nome': cor  # Aqui tem a cor da célula do nome, ex: 'FFFF0000'
                }
                escala.append(item)

        logging.info(f"Processadas {len(escala)} linhas.")
        return escala

    except Exception as e:
        logging.error(f"Erro ao processar Excel: {e}", exc_info=True)
        return []
